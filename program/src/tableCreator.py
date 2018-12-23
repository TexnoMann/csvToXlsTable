from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
import string

class XLSTable():
    def __init__(self, filename):
        self.__filename=filename
        self.__wb=Workbook()

        self.__standartFont = Font(name='Calibri',size=12,bold=False,italic=False, vertAlign=None,underline='none',strike=False)
        self.__standartFont_bold = Font(name='Calibri',size=12,bold=True,italic=False, vertAlign=None,underline='none',strike=False)

        sheet1=self.__wb.active
        sheet1.title="Сводная страница"
        sheet2=self.__wb.create_sheet(title='Расписание')
        sheet3=self.__wb.create_sheet(title='Условные обозначения')
        self.__sheets=[sheet1,sheet2,sheet3]

    def __makeHead(self,list,numbersheet):
        self.__writeListInRow(list,1, numbersheet)

    def __writeListInRow(self,list, numberRow, numbersheet):
        for i in range(0,len(list)):
            self.__sheets[numbersheet].cell(row=numberRow,column=i+1).value=list[i]
            self.__sheets[numbersheet].cell(row=numberRow,column=i+1).font=self.__standartFont

    def writeRectDataInTable(self,list,startrow,startcol, numbersheet):
        for i in range(0,len(list)):
            for j in range(0,len(list[i])):
                self.__sheets[numbersheet].cell(row=i+startrow,column=j+startcol).value=list[i][j]
                self.__sheets[numbersheet].cell(row=i+startrow,column=j+startcol).font=self.__standartFont

    def saveInFile(self):
        self.__wb.save(self.__filename)

    def setFontStyle(self, font, nrow, ncolumn, numbersheet):
        self.__sheets[numbersheet].cell(row=nrow, column=ncolumn).font=font

    def autoFitDimmensionsCollumn(self, numbersheet,ignorecolumns=[]):
        for col in self.__sheets[numbersheet].columns:
            max_length = 0
            column = col[0].column
            if not column in ignorecolumns:
                for cell in col:
                    if cell.coordinate in self.__sheets[numbersheet].merged_cells:
                        if len(str(cell.value)) > max_length:
                            max_length=len(str(cell.value))
                            adjusted_width = (max_length + 2) * 1.2*0.5
                            self.__sheets[numbersheet].column_dimensions[column].width = adjusted_width
                            continue
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                self.__sheets[numbersheet].column_dimensions[column].width = adjusted_width

    def makeBoldFontforColumn(self, numbersheet, numbercolumns):
        for i in range(0,self.__sheets[numbersheet].max_column):
            for j in numbercolumns:
                self.__sheets[numbersheet].cell(row=i+1,column=j).font=self.__standartFont_bold

    def makeBoldFontforRow(self,numbersheet, numberrows):
        for i in range(0,1000):
            for j in numberrows:
                self.__sheets[numbersheet].cell(row=j,column=i+1).font=self.__standartFont_bold

    def initFillTables(self):
        template_sheet2=[['Корпус','Вид/Пример'],
                ['Кронверкский','285 Кр'],
                ['Ломоносова','1222 Лом'],
                ['Биржевая','203 Бир'],
                ['Гривцова','409 Гр'],
                ['Чайковского',	'305 Чай'],
                ['Гастелло','403 Гаст'],
                ['Хрустальная','39 Хр']]
        template_sheet0=[['','','','ФАКТ**','',''],
                        ['Дисциплина','Тип занятия','ПЛАН*']]

        self.__sheets[2].merge_cells('A1:B1')
        self.__sheets[2].merge_cells('A2:B2')
        self.__sheets[2].merge_cells('A3:B3')

        self.__sheets[0].merge_cells('A1:E1')
        self.__makeHead(['Обозначение номеров аудиторий'],2)
        self.__makeHead(['Дата','Номер недели','День недели','Номер пары','Дисциплина','Вид занятия','Преподаватель','Аудитория','Группы'],1)
        self.__makeHead(['Шапка РУП'],0)

        self.__writeListInRow(['В целях единообразия устанивливается единая система записи аудиторий.'],2,2)
        self.__writeListInRow(['Сначала ставится номер аудитории,\n затем пробел, затем сокращеное наименование корпуса без точки.'],3,2)
        self.writeRectDataInTable(template_sheet2,5,1,2)
        self.writeRectDataInTable(template_sheet0,2,1,0)

        self.__sheets[2].row_dimensions[3].height = 30
        self.__sheets[0].column_dimensions['A'].width = 20
        self.__sheets[0].column_dimensions['B'].width = 20

        self.makeBoldFontforRow(0,[1,2,3])
        self.makeBoldFontforRow(1,[1])
        self.makeBoldFontforRow(2,[1,5])

    def getMaxTableSize(self, numbersheet):
        maxR=0
        maxC=0
        for i in range(1,self.__sheets[numbersheet].max_row+1):
            for j in range(1,self.__sheets[numbersheet].max_column+1):
                p=self.__sheets[numbersheet].cell(row=i,column=j).value
                if(p!=None):
                    maxR=i
                    maxC=j
        return [maxR,maxC]

    def getHeadTable(self,numbersheet):
        countC=self.getMaxTableSize(numbersheet)[1]
        return self.getSquareDataOfTable(numbersheet,[1,1],[1,countC])[0]

    def getSquareDataOfTable(self,numbersheet,startP,finishP):
        data=[]
        for i in range(startP[0],finishP[0]+1):
            row=[]
            for j in range(startP[1], finishP[1]+1):
                p=self.__sheets[numbersheet].cell(row=i,column=j).value
                row.append(p)
            data.append(copy(row))
        return copy(data)

    def getColumnNumberwithHeadName(self,numbersheet, colName):
        head=self.getHeadTable(numbersheet)
        return head.index(colName)+1
